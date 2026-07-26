"""
Excel export using OpenPyXL — multi-sheet workbook with real formatting
(header styling, column widths, number formats), not just a raw dataframe dump.
"""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def _style_header(ws, n_cols):
    header_fill = PatternFill(start_color="171A21", end_color="171A21", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, df):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 40)


def _write_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _style_header(ws, len(df.columns))
    _autofit_columns(ws, df)
    ws.freeze_panes = "A2"


def build_excel_report(kpis: dict, sales_df: pd.DataFrame, segments_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # KPI summary sheet
    kpi_df = pd.DataFrame([
        {"Metric": "Revenue", "Value": f"${kpis['revenue']:,.2f}"},
        {"Metric": "Orders", "Value": kpis["orders"]},
        {"Metric": "Customers", "Value": kpis["customers"]},
        {"Metric": "Profit (est.)", "Value": f"${kpis['profit']:,.2f}"},
        {"Metric": "MoM Growth", "Value": f"{kpis['growth']:.1f}%"},
    ])
    _write_sheet(wb, "Executive Summary", kpi_df)

    _write_sheet(wb, "Sales Detail", sales_df)
    _write_sheet(wb, "Customer Segments", segments_df)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()